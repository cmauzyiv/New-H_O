import pygame
import sys
import os
import constance as c
from background import BG1, BG2, BG3, BG4
from ship import Ship
from bg_spawner import P1BgSpawner, P2BgSpawner, P3BgSpawner, P4BgSpawner
from enemy_spawner import EnemySpawner
from p2enemy_spawner import p2EnemySpawner
from p3enemy_spawner import p3EnemySpawner
from particle_spawner import ParticleSpawner
from explosion_spawner import ExplosionSpawner
import openpyxl
from openpyxl import Workbook, load_workbook
from Asteroid_HuD import Asteroid_HUD



pygame.mixer.pre_init(44100, -16, 2, 512)
pygame.init()
pygame.mixer.init()
pygame.font.init()
##pygame.mixer.music.load("theme_music.ogg")
##pygame.mixer.music.set_volume(.5)
##pygame.mixer.music.play(loops=True)
##stream_path = r'C:\\Users\\mauzyc\\OneDrive\\Python Programs\\Multiple Schedule Study\\Main Files 1\\1.stream.txt'
##data_path = r'C:\\Users\\mauzyc\\OneDrive\\Python Programs\\Multiple Schedule Study\\Main Files 1\\1.data.txt'
fw = open('1.data.txt', 'w')
fw_stream = open('1.stream.txt', 'w')
fw_seconds = open('seconds.txt', 'w')


## power ups for each color gem
## ahve it count how many of each color

## if player.rect.x >= enemy_group1.rect.x:
    ##player.rect.x *<1 - so if player was to right of enemy, he bounces right


phase = 1
trial_type = 0
trial_num = 1
experiment_type = 1
thin_pbx = True
thin_discr = False
one_min_sessions = True
two_min_sessions = False
long_baseline = True
alternative_activities = True
session = 1
tg_1 = 0
alt_1 = 0
sd_alt_1 = 0
sdelta_alt_1 = 0
ms_tg_1 = 0
score = 0
asteroids = 0
bsl_target = 0
phase_target = 0
phase_alt = 0
phase_score = 0
phase_asteroids = 0
barrel_rolls = 0
spins = 0
coin_boosters = 0
turbo = 0
blue_gem_token = 0
yellow_gem_token = 0
red_gem_token = 0
crash = 0








def main():
    global phase
    global trial_type
    global trial_num
    global session
    global tg_1
    global alt_1
    global score
    global asteroids
    global sd_alt_1
    global sdelta_alt_1
    global ms_tg_1
    global bsl_target
    global phase_alt
    global phase_target
    global phase_score
    global phase_asteroids
    all_asteroids = False
    run = True
    fps = 60
    clock = pygame.time.Clock()
    clock1 = pygame.time.Clock()
    delta_time = clock1.tick(60) / 1000
    level = 0
    display = pygame.display.set_mode(c.display_size) ##(c.display_size, pygame.FULLSCREEN | pygame.HWACCEL)
    black = (0, 0, 0)
    display.fill(black)
    discrimination = 0
    trial_timer = 7200
    phase_timer = 36000
    long_sdelta_timer = 18000
    long_sd1_timer = 9000
    long_sd2_timer = 9000
    short_sdelta_timer = 3600
    short_sd_timer = 3600
    shorter_sdelta_timer = 1800
    shorter_sd_timer = 1800
    thinning_step = 1
    time_count = 0
    seconds_count = 0
    stream = []
    number_shots = []
    seconds_timer = 60
    barrel_rolls = 0
    spins = 0
    coin_boosters = 0
    turbo = 0
    blue_gem_token = 0
    yellow_gem_token = 0
    red_gem_token = 0
    crash = 0
    flame_change = 0
    if not long_baseline:
        wb = load_workbook('1.graph.xlsx')
        ws = wb['Data Entry 1']
        workbook = '1.graph.xlsx'
    if long_baseline:
        wb = load_workbook('1.long_bsl_graph.xlsx')
        ws = wb['Data Entry 1']
        workbook = '1.long_bsl_graph.xlsx'

    ##ws['B8'].value = 2
    ##wb.save("graph.xlsx")
    ##print(ws['B8'].value)










## Groups/Spawners
    player = Ship()
    sprite_group = pygame.sprite.Group()
    sprite_group.add(player)
    ##shield_spawner = ShieldSpawner()
    particle_spawner = ParticleSpawner()
    p2enemy_spawner = p2EnemySpawner()
    enemy_spawner = EnemySpawner()
    p3enemy_spawner = p3EnemySpawner()
    explosion_spawner = ExplosionSpawner()
    p1bg_spawner = P1BgSpawner()
    p2bg_spawner = P2BgSpawner()
    p3bg_spawner = P3BgSpawner()
    p4bg_spawner = P4BgSpawner()



    def set_timers():
        nonlocal short_sd_timer
        nonlocal short_sdelta_timer
        if thinning_step == 1:
            short_sd_timer = 15
            short_sdelta_timer = 5
        elif thinning_step == 2:
            short_sd_timer = 10
            short_sdelta_timer = 10
        elif thinning_step == 3:
            short_sd_timer = 5
            short_sdelta_timer = 15

    set_timers()

    def record_data():
        global alt_1
        global tg_1
        global session
        global phase
        global asteroids
        global score
        global trial_type
        global trial_num
        global sd_alt_1
        global sdelta_alt_1
        nonlocal barrel_rolls
        nonlocal crash
        nonlocal blue_gem_token
        nonlocal yellow_gem_token
        nonlocal red_gem_token
        nonlocal spins
        nonlocal coin_boosters
        nonlocal flame_change
        nonlocal turbo

        fw.write('\n')
        fw.write('phase = ')
        if phase == 1:
            fw.write('Baseline')
            fw.write('\n')
        elif phase == 2:
            fw.write('Treatment')
            fw.write('\n')
        elif phase == 3:
            fw.write('Sd')
            fw.write('\n')
        elif phase == 4:
            fw.write('Extinction')
            fw.write('\n')
        fw.write('trial = ')
        fw.write(str(trial_num))
        fw.write('\n')
        fw.write('target = ')
        fw.write('\n')
        fw.write(str(tg_1))
        fw.write('\n')
        if phase >= 2:
            fw.write('alternative = ')
            fw.write('\n')
            fw.write(str(alt_1))
            fw.write('\n')
        fw.write('asteroids hit =')
        fw.write('\n')
        fw.write(str(asteroids))
        fw.write('\n')
        fw.write('score = ')
        fw.write('\n')
        fw.write(str(score))
        fw.write('\n')
        fw.write('power ups:')
        fw.write('\n')
        fw.write('barrel rolls = ')
        fw.write(str(barrel_rolls))
        fw.write(', spins = ')
        fw.write(str(spins))
        fw.write(', speed boost = ')
        fw.write(str(turbo))
        fw.write(', coin boost = ')
        fw.write(str(coin_boosters))
        fw.write(', flame change = ')
        fw.write(str(flame_change))
        fw.write('\n')
        fw.write('token changes = ')
        fw.write('\n')
        fw.write('red = ')
        fw.write(str(red_gem_token))
        fw.write(', blue = ')
        fw.write(str(blue_gem_token))
        fw.write(', yellow = ')
        fw.write(str(yellow_gem_token))
        fw.write('\n')
        fw.write('collisions = ')
        fw.write(str(crash))
        fw.write('\n')
        if trial_type == 1:
            fw.write(str(discrimination))

        alt_1 = 0
        tg_1 = 0
        asteroids = 0
        barrel_rolls = 0
        spins = 0
        coin_boosters = 0
        turbo = 0
        blue_gem_token = 0
        yellow_gem_token = 0
        red_gem_token = 0
        crash = 0
        flame_change = 0




    def record_3min_data():
        global alt_1
        global tg_1
        global session
        global phase
        global asteroids
        global score
        global trial_type
        global trial_num
        global sd_alt_1
        global sdelta_alt_1
        global phase_target
        global phase_alt
        global phase_score
        global phase_asteroids
        fw.write('\n')
        fw.write('phase = ')
        if phase == 1:
            fw.write('Baseline')
            fw.write('\n')
        elif phase == 2:
            fw.write('Treatment')
            fw.write('\n')
        elif phase == 3:
            fw.write('Sd')
            fw.write('\n')
        elif phase == 4:
            fw.write('Extinction')
            fw.write('\n')
        fw.write('trial = ')
        fw.write(str(trial_num))
        fw.write('\n')
        fw.write('target = ')
        fw.write('\n')
        fw.write(str(phase_target))
        fw.write('\n')
        if phase >= 2:
            fw.write('alternative = ')
            fw.write('\n')
            fw.write(str(phase_alt))
            fw.write('\n')
        fw.write('asteroids hit =')
        fw.write('\n')
        fw.write(str(phase_asteroids))
        fw.write('\n')
        fw.write('score = ')
        fw.write('\n')
        fw.write(str(phase_score))
        fw.write('\n')
        if trial_type == 1:
            fw.write(str(discrimination))

        phase_target = 0
        phase_alt = 0
        phase_asteroids = 0

    def graph_data():
        global alt_1
        global tg_1
        global session
        global phase
        global asteroids
        global score
        global trial_type
        global trial_num
        global sd_alt_1
        global sdelta_alt_1
        nonlocal barrel_rolls
        nonlocal crash
        nonlocal blue_gem_token
        nonlocal yellow_gem_token
        nonlocal red_gem_token
        nonlocal spins
        nonlocal coin_boosters
        nonlocal flame_change
        nonlocal turbo
        if not long_baseline:
            if trial_num == 1:
                ws['C4'].value = phase
                ws['D4'].value = tg_1
                ws['E4'].value = alt_1
                ws['F4'].value = asteroids
                ws['G4'].value = score
                ws['H4'].value = barrel_rolls
                ws['I4'].value = spins
                ws['J4'].value = turbo
                ws['K4'].value = coin_boosters
                ws['L4'].value = flame_change
                ws['M4'].value = red_gem_token
                ws['N4'].value = blue_gem_token
                ws['O4'].value = yellow_gem_token
                ws['P4'].value = crash
                wb.save(workbook)
            if trial_num == 2:
                ws['C5'].value = phase
                ws['D5'].value = tg_1
                ws['E5'].value = alt_1
                ws['F5'].value = asteroids
                ws['G5'].value = score
                ws['H5'].value = barrel_rolls
                ws['I5'].value = spins
                ws['J5'].value = turbo
                ws['K5'].value = coin_boosters
                ws['L5'].value = flame_change
                ws['M5'].value = red_gem_token
                ws['N5'].value = blue_gem_token
                ws['O5'].value = yellow_gem_token
                ws['P5'].value = crash
                wb.save(workbook)
            if trial_num == 3:
                ws['C6'].value = phase
                ws['D6'].value = tg_1
                ws['E6'].value = alt_1
                ws['F6'].value = asteroids
                ws['G6'].value = score
                ws['H6'].value = barrel_rolls
                ws['I6'].value = spins
                ws['J6'].value = turbo
                ws['K6'].value = coin_boosters
                ws['L6'].value = flame_change
                ws['M6'].value = red_gem_token
                ws['N6'].value = blue_gem_token
                ws['O6'].value = yellow_gem_token
                ws['P6'].value = crash
                wb.save(workbook)
            if trial_num == 4:
                ws['C7'].value = phase
                ws['D7'].value = tg_1
                ws['E7'].value = alt_1
                ws['F7'].value = asteroids
                ws['G7'].value = score
                ws['H7'].value = barrel_rolls
                ws['I7'].value = spins
                ws['J7'].value = turbo
                ws['K7'].value = coin_boosters
                ws['L7'].value = flame_change
                ws['M7'].value = red_gem_token
                ws['N7'].value = blue_gem_token
                ws['O7'].value = yellow_gem_token
                ws['P7'].value = crash
                wb.save(workbook)
            if trial_num == 5:
                ws['C8'].value = phase
                ws['D8'].value = tg_1
                ws['E8'].value = alt_1
                ws['F8'].value = asteroids
                ws['G8'].value = score
                ws['H8'].value = barrel_rolls
                ws['I8'].value = spins
                ws['J8'].value = turbo
                ws['K8'].value = coin_boosters
                ws['L8'].value = flame_change
                ws['M8'].value = red_gem_token
                ws['N8'].value = blue_gem_token
                ws['O8'].value = yellow_gem_token
                ws['P8'].value = crash
                wb.save(workbook)
            if trial_num == 6:
                ws['C9'].value = phase
                ws['D9'].value = tg_1
                ws['E9'].value = alt_1
                ws['F9'].value = asteroids
                ws['G9'].value = score
                ws['H9'].value = barrel_rolls
                ws['I9'].value = spins
                ws['J9'].value = turbo
                ws['K9'].value = coin_boosters
                ws['L9'].value = flame_change
                ws['M9'].value = red_gem_token
                ws['N9'].value = blue_gem_token
                ws['O9'].value = yellow_gem_token
                ws['P9'].value = crash
                wb.save(workbook)
            if trial_num == 7:
                ws['C10'].value = phase
                ws['D10'].value = tg_1
                ws['E10'].value = alt_1
                ws['F10'].value = asteroids
                ws['G10'].value = score
                ws['H10'].value = barrel_rolls
                ws['I10'].value = spins
                ws['J10'].value = turbo
                ws['K10'].value = coin_boosters
                ws['L10'].value = flame_change
                ws['M10'].value = red_gem_token
                ws['N10'].value = blue_gem_token
                ws['O10'].value = yellow_gem_token
                ws['P10'].value = crash
                wb.save(workbook)
            if trial_num == 8:
                ws['C11'].value = phase
                ws['D11'].value = tg_1
                ws['E11'].value = alt_1
                ws['F11'].value = asteroids
                ws['G11'].value = score
                ws['H11'].value = barrel_rolls
                ws['I11'].value = spins
                ws['J11'].value = turbo
                ws['K11'].value = coin_boosters
                ws['L11'].value = flame_change
                ws['M11'].value = red_gem_token
                ws['N11'].value = blue_gem_token
                ws['O11'].value = yellow_gem_token
                ws['P11'].value = crash
                wb.save(workbook)
            if trial_num == 9:
                ws['C12'].value = phase
                ws['D12'].value = tg_1
                ws['E12'].value = alt_1
                ws['F12'].value = asteroids
                ws['G12'].value = score
                ws['H12'].value = barrel_rolls
                ws['I12'].value = spins
                ws['J12'].value = turbo
                ws['K12'].value = coin_boosters
                ws['L12'].value = flame_change
                ws['M12'].value = red_gem_token
                ws['N12'].value = blue_gem_token
                ws['O12'].value = yellow_gem_token
                ws['P12'].value = crash
                wb.save(workbook)
            if trial_num == 10:
                ws['C13'].value = phase
                ws['D13'].value = tg_1
                ws['E13'].value = alt_1
                ws['F13'].value = asteroids
                ws['G13'].value = score
                ws['H13'].value = barrel_rolls
                ws['I13'].value = spins
                ws['J13'].value = turbo
                ws['K13'].value = coin_boosters
                ws['L13'].value = flame_change
                ws['M13'].value = red_gem_token
                ws['N13'].value = blue_gem_token
                ws['O13'].value = yellow_gem_token
                ws['P13'].value = crash
                wb.save(workbook)
            if trial_num == 11:
                ws['C14'].value = phase
                ws['D14'].value = tg_1
                ws['E14'].value = alt_1
                ws['F14'].value = asteroids
                ws['G14'].value = score
                ws['H14'].value = barrel_rolls
                ws['I14'].value = spins
                ws['J14'].value = turbo
                ws['K14'].value = coin_boosters
                ws['L14'].value = flame_change
                ws['M14'].value = red_gem_token
                ws['N14'].value = blue_gem_token
                ws['O14'].value = yellow_gem_token
                ws['P14'].value = crash
                wb.save(workbook)
        if long_baseline:
            if trial_num == 1:
                ws['C4'].value = phase
                ws['D4'].value = tg_1
                ws['E4'].value = alt_1
                ws['F4'].value = asteroids
                ws['G4'].value = score
                ws['H4'].value = barrel_rolls
                ws['I4'].value = spins
                ws['J4'].value = turbo
                ws['K4'].value = coin_boosters
                ws['L4'].value = flame_change
                ws['M4'].value = red_gem_token
                ws['N4'].value = blue_gem_token
                ws['O4'].value = yellow_gem_token
                ws['P4'].value = crash
                wb.save(workbook)
            if trial_num == 2:
                ws['C5'].value = phase
                ws['D5'].value = tg_1
                ws['E5'].value = alt_1
                ws['F5'].value = asteroids
                ws['G5'].value = score
                ws['H5'].value = barrel_rolls
                ws['I5'].value = spins
                ws['J5'].value = turbo
                ws['K5'].value = coin_boosters
                ws['L5'].value = flame_change
                ws['M5'].value = red_gem_token
                ws['N5'].value = blue_gem_token
                ws['O5'].value = yellow_gem_token
                ws['P5'].value = crash
                wb.save(workbook)
            if trial_num == 3:
                ws['C6'].value = phase
                ws['D6'].value = tg_1
                ws['E6'].value = alt_1
                ws['F6'].value = asteroids
                ws['G6'].value = score
                ws['H6'].value = barrel_rolls
                ws['I6'].value = spins
                ws['J6'].value = turbo
                ws['K6'].value = coin_boosters
                ws['L6'].value = flame_change
                ws['M6'].value = red_gem_token
                ws['N6'].value = blue_gem_token
                ws['O6'].value = yellow_gem_token
                ws['P6'].value = crash
                wb.save(workbook)
            if trial_num == 4:
                ws['C7'].value = phase
                ws['D7'].value = tg_1
                ws['E7'].value = alt_1
                ws['F7'].value = asteroids
                ws['G7'].value = score
                ws['H7'].value = barrel_rolls
                ws['I7'].value = spins
                ws['J7'].value = turbo
                ws['K7'].value = coin_boosters
                ws['L7'].value = flame_change
                ws['M7'].value = red_gem_token
                ws['N7'].value = blue_gem_token
                ws['O7'].value = yellow_gem_token
                ws['P7'].value = crash
                wb.save(workbook)
            if trial_num == 5:
                ws['C8'].value = phase
                ws['D8'].value = tg_1
                ws['E8'].value = alt_1
                ws['F8'].value = asteroids
                ws['G8'].value = score
                ws['H8'].value = barrel_rolls
                ws['I8'].value = spins
                ws['J8'].value = turbo
                ws['K8'].value = coin_boosters
                ws['L8'].value = flame_change
                ws['M8'].value = red_gem_token
                ws['N8'].value = blue_gem_token
                ws['O8'].value = yellow_gem_token
                ws['P8'].value = crash
                wb.save(workbook)
            if trial_num == 6:
                ws['C9'].value = phase
                ws['D9'].value = tg_1
                ws['E9'].value = alt_1
                ws['F9'].value = asteroids
                ws['G9'].value = score
                ws['H9'].value = barrel_rolls
                ws['I9'].value = spins
                ws['J9'].value = turbo
                ws['K9'].value = coin_boosters
                ws['L9'].value = flame_change
                ws['M9'].value = red_gem_token
                ws['N9'].value = blue_gem_token
                ws['O9'].value = yellow_gem_token
                ws['P9'].value = crash
                wb.save(workbook)
            if trial_num == 7:
                ws['C10'].value = phase
                ws['D10'].value = tg_1
                ws['E10'].value = alt_1
                ws['F10'].value = asteroids
                ws['G10'].value = score
                ws['H10'].value = barrel_rolls
                ws['I10'].value = spins
                ws['J10'].value = turbo
                ws['K10'].value = coin_boosters
                ws['L10'].value = flame_change
                ws['M10'].value = red_gem_token
                ws['N10'].value = blue_gem_token
                ws['O10'].value = yellow_gem_token
                ws['P10'].value = crash
                wb.save(workbook)
            if trial_num == 8:
                ws['C11'].value = phase
                ws['D11'].value = tg_1
                ws['E11'].value = alt_1
                ws['F11'].value = asteroids
                ws['G11'].value = score
                ws['H11'].value = barrel_rolls
                ws['I11'].value = spins
                ws['J11'].value = turbo
                ws['K11'].value = coin_boosters
                ws['L11'].value = flame_change
                ws['M11'].value = red_gem_token
                ws['N11'].value = blue_gem_token
                ws['O11'].value = yellow_gem_token
                ws['P11'].value = crash
                wb.save(workbook)
            if trial_num == 9:
                ws['C12'].value = phase
                ws['D12'].value = tg_1
                ws['E12'].value = alt_1
                ws['F12'].value = asteroids
                ws['G12'].value = score
                ws['H12'].value = barrel_rolls
                ws['I12'].value = spins
                ws['J12'].value = turbo
                ws['K12'].value = coin_boosters
                ws['L12'].value = flame_change
                ws['M12'].value = red_gem_token
                ws['N12'].value = blue_gem_token
                ws['O12'].value = yellow_gem_token
                ws['P12'].value = crash
                wb.save(workbook)
            if trial_num == 10:
                ws['C13'].value = phase
                ws['D13'].value = tg_1
                ws['E13'].value = alt_1
                ws['F13'].value = asteroids
                ws['G13'].value = score
                ws['H13'].value = barrel_rolls
                ws['I13'].value = spins
                ws['J13'].value = turbo
                ws['K13'].value = coin_boosters
                ws['L13'].value = flame_change
                ws['M13'].value = red_gem_token
                ws['N13'].value = blue_gem_token
                ws['O13'].value = yellow_gem_token
                ws['P13'].value = crash
                wb.save(workbook)
            if trial_num == 11:
                ws['C14'].value = phase
                ws['D14'].value = tg_1
                ws['E14'].value = alt_1
                ws['F14'].value = asteroids
                ws['G14'].value = score
                ws['H14'].value = barrel_rolls
                ws['I14'].value = spins
                ws['J14'].value = turbo
                ws['K14'].value = coin_boosters
                ws['L14'].value = flame_change
                ws['M14'].value = red_gem_token
                ws['N14'].value = blue_gem_token
                ws['O14'].value = yellow_gem_token
                ws['P14'].value = crash
                wb.save(workbook)

    def update_stream():
        nonlocal seconds_count
        nonlocal stream
        nonlocal number_shots
        fw_stream.write(str(seconds_count))
        fw_stream.write(str(stream))
        fw_stream.write('\n')
        stream = []


    def update_groups():
        sprite_group.update()
        particle_spawner.update()
        explosion_spawner.update()
        if phase == 1:
            p1bg_spawner.update()
        elif phase == 2:
            p2bg_spawner.update()
        elif phase == 3:
            if trial_type == 0:
                p3bg_spawner.update()
            elif trial_type == 1:
                p4bg_spawner.update()
        elif phase == 4:
            p4bg_spawner.update()
        enemy_spawner.update()
        if phase == 5:
            game_over()



    def update_display():
        particle_spawner.particle_group.draw(display)
        player.hud_group.draw(display)
        player.hud.score_group.draw(display)
        if experiment_type == 2:
            player.hud.level_group.draw(display)
        sprite_group.draw(display)
        explosion_spawner.p1asteroidexplosions_group.draw(display)
        explosion_spawner.p2asteroidexplosions_group.draw(display)
        explosion_spawner.sdasteroidexplosions_group.draw(display)
        explosion_spawner.sdeltaasteroidexplosions_group.draw(display)
        explosion_spawner.shipexplosions_group.draw(display)
        ##shield_spawner.shipshield_group.draw(display)
        player.roll_effect_group.draw(display)
        player.lasers.draw(display)
        player.lmuzzle_group.draw(display)
        player.rmuzzle_group.draw(display)
        player.lengine_group.draw(display)
        player.rengine_group.draw(display)
        player.ShipShield_group.draw(display)
        player.lflame_group.draw(display)
        player.rflame_group.draw(display)
        if alternative_activities:
            player.asteroid_hud.yscore_group.draw(display)
            player.asteroid_hud.bscore_group.draw(display)
            player.asteroid_hud.rscore_group.draw(display)
            enemy_spawner.spin_group.draw(display)
            enemy_spawner.enemy_group3.draw(display)
            enemy_spawner.speed_group.draw(display)
            enemy_spawner.max_gems_group.draw(display)
            enemy_spawner.roll_group.draw(display)
            enemy_spawner.asteroid_group.draw(display)
            enemy_spawner.yellow_icon_group.draw(display)
            enemy_spawner.blue_icon_group.draw(display)
            enemy_spawner.red_icon_group.draw(display)
            enemy_spawner.red_flame_group.draw(display)
            enemy_spawner.blue_flame_group.draw(display)
        pygame.display.update()

## Phase specific
        if phase == 1:
            p1bg_spawner.bg_group.draw(display)
            p1bg_spawner.stars_group.draw(display)
            enemy_spawner.enemy_group1.draw(display)  ## the enemy spawner accesses the enemy group and draws it on display
            global enemy
            for enemy in enemy_spawner.enemy_group1:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
                enemy.particle_group.draw(display)
        elif phase == 2:
            p2bg_spawner.bg_group.draw(display)
            p2bg_spawner.stars_group.draw(display)
            enemy_spawner.enemy_group1.draw(display)  ## the enemy spawner accesses the enemy group and draws it on display
            for enemy in enemy_spawner.enemy_group1:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
                enemy.particle_group.draw(display)
            enemy_spawner.enemy_group2.draw(display)
            for enemy in enemy_spawner.enemy_group2:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
                enemy.particle_group.draw(display)
        if phase == 3:
            if trial_type == 0:
                p3bg_spawner.bg_group.draw(display)
                p3bg_spawner.stars_group.draw(display)
            elif trial_type == 1:
                p4bg_spawner.bg_group.draw(display)
                p4bg_spawner.stars_group.draw(display)
            enemy_spawner.enemy_group1.draw(display)  ## the enemy spawner accesses the enemy group and draws it on display
            for enemy in enemy_spawner.enemy_group1:
                enemy.engine_group.draw(display)
            enemy_spawner.enemy_group2.draw(display)
            for enemy in enemy_spawner.enemy_group2:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
                enemy.particle_group.draw(display)
        if phase == 4:
            p4bg_spawner.bg_group.draw(display)
            p4bg_spawner.stars_group.draw(display)
            enemy_spawner.enemy_group1.draw(display)  ## the enemy spawner accesses the enemy group and draws it on display
            for enemy in enemy_spawner.enemy_group1:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
            enemy_spawner.enemy_group2.draw(display)
            for enemy in enemy_spawner.enemy_group2:
                enemy.engine_group.draw(display)
                enemy.ShipShield_group.draw(display)
        if phase == 5:
            game_over()




    def calculate_discrimination():
        global phase
        global trial_type
        global sd_alt_1
        global sdelta_alt_1
        nonlocal thinning_step
        nonlocal discrimination
        if sd_alt_1 == 0:
            trial_type = 0
        else:
            discrimination = (sd_alt_1 / (sd_alt_1 + sdelta_alt_1))
            print(str(discrimination))
            high_discrimination = (sd_alt_1 / (sd_alt_1 + sdelta_alt_1)) >= .8
            low_discrimination = (sd_alt_1 / (sd_alt_1 + sdelta_alt_1)) < .8
            if high_discrimination:
                if thinning_step < 3:
                    thinning_step += 1
                    trial_type = 0
                    player.hud.level.value += 1
                else:
                    game_over()
                print("high_discrimination")
            elif low_discrimination:
                trial_type = 0
                print("low_discrimination")

    def calculate_reduction():
        global tg_1
        global bsl_target
        global trial_type
        nonlocal thinning_step
        bsl_rate = bsl_target / 3
        reduction = bsl_rate * .2
        print(reduction)
        if tg_1 <= reduction:
            if thinning_step < 3:
                thinning_step += 1
                trial_type = 0
                player.hud.level.value += 1
            else:
                game_over()
        else:
            trial_type = 0


## Main loop
    while run:
        update_groups()
        update_display()
        time_count += 1
        clock.tick(fps)
        seconds_timer -= 1
        if experiment_type == 1:
            if seconds_timer == 0:
                seconds_count += 1
                update_stream()
                seconds_timer = 60
                if seconds_count == 600:
                    seconds_count = 0
            if one_min_sessions:
                if seconds_count == 60:
                    graph_data()
                    record_data()
                    if not long_baseline:
                        if trial_num == 3:
                            phase = 2
                            enemy_spawner.phase = 2
                        elif trial_num == 6:
                            phase = 4
                            enemy_spawner.phase = 4
                        elif trial_num == 9:
                            phase = 5
                        trial_num += 1
                        seconds_count = 0
                    if long_baseline:
                        if trial_num == 4:
                            phase = 2
                            enemy_spawner.phase = 2
                        elif trial_num == 7:
                            phase = 4
                            enemy_spawner.phase = 4
                        elif trial_num == 10:
                            phase = 5
                        trial_num += 1
                        seconds_count = 0
            elif two_min_sessions:
                if seconds_count == 120:
                    graph_data()
                    record_data()
                    if trial_num == 3:
                        phase = 2
                        enemy_spawner.phase = 2
                    elif trial_num == 6:
                        phase = 4
                        enemy_spawner.phase = 4
                    elif trial_num == 9:
                        phase = 5
                    trial_num += 1
                    seconds_count = 0
        elif experiment_type == 2:
            if seconds_timer == 0:
                seconds_count += 1
                if trial_type == 0:
                    short_sd_timer -= 1
                if trial_type == 1:
                    short_sdelta_timer -= 1
                update_stream()
                seconds_timer = 60
                if seconds_count == 600:
                    seconds_count = 0
            if seconds_count == 60:
                record_data()
                if trial_num == 3:
                    phase = 2
                    enemy_spawner.phase = 2
                    player.hud.level.value += 1
                elif trial_num == 6:
                    phase = 3
                    trial_type = 0
                    set_timers()
                    player.hud.level.value += 1
                elif trial_num >= 7:
                    if thin_discr:
                        calculate_discrimination()
                    elif thin_pbx:
                        calculate_reduction()
                    set_timers()
                trial_num += 1
                seconds_count = 0
            ## Setting up sd and sdelta
            if phase == 3:
                if short_sd_timer == 0:
                    trial_type = 1
                if short_sdelta_timer == 0:
                    trial_type = 0
                    set_timers()





## Event handler
        for event in pygame.event.get():
            if event.type == pygame.QUIT or (event.type == pygame.KEYDOWN and event.key == pygame.K_ESCAPE):
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_a:
                    player.vel_x = -player.speed
                    player.is_moving_x = True
                    ##if player.kickbacked:
                       ## player.is_moving = True
                       ## player.speed *= - 1
                    if not player.spinning and not player.rolling_left and not player.rolling_right:
                        player.turning_left = True
                        player.anim_index = 0
                    if player.vel_y == 0:
                        player.rev_engine()

                    ##player.ShipShield_group.vel_x = -shields.speed
                elif event.key == pygame.K_d:
                    player.moving_x = True
                    player.vel_x = player.speed
                    if not player.spinning and not player.rolling_left and not player.rolling_right:
                        player.turning_right = True
                        player.anim_index = 0
                    if player.vel_y == 0:
                        player.rev_engine()
                    ##player.ShipShield_group.vel_x = shields.speed
                elif event.key == pygame.K_w:
                    player.vel_y = -player.speed
                    player.is_moving_y = True
                    if player.vel_x == 0:
                        player.rev_engine()
                    ##player.ShipShield_group.vel_y = -shields.speed
                elif event.key == pygame.K_s:
                    player.vel_y = player.speed
                    player.is_moving_y = True
                    if player.vel_x == 0:
                        player.rev_engine()
                    ##player.ShipShield_group.vel_y = shields.speed
                if event.key == pygame.K_SPACE:
                    player.shoot()
                    stream.append(9)
            if event.type == pygame.MOUSEBUTTONDOWN:
                player.shoot()
                stream.append(9)
            if event.type == pygame.KEYUP:
                if event.key == pygame.K_a:
                    player.vel_x = 0
                    player.is_moving_x = False
                    if not player.spinning and not player.rolling_left and not player.rolling_right:
                        player.go_straight()
                        player.turning_left = False
                    if player.vel_y == 0:
                        player.engine_die()
                    ##shields.vel_x = 0
                elif event.key == pygame.K_d:
                    player.vel_x = 0
                    player.is_moving_x = False
                    if not player.spinning and not player.rolling_left and not player.rolling_right:
                        player.go_straight()
                        player.turning_right = False
                    if player.vel_y == 0:
                        player.engine_die()
                    ##shields.vel_x = 0
                elif event.key == pygame.K_s:
                    player.vel_y = 0
                    player.is_moving_y = False
                    if player.vel_x == 0:
                        player.engine_die()
                    ##shields.vel_y = 0
                elif event.key == pygame.K_w:
                    player.vel_y = 0
                    player.is_moving_y = False
                    if player.vel_x == 0:
                        player.engine_die()
                    ##shields.vel_x = 0



        ##collisions:

        collided = pygame.sprite.groupcollide(player.lasers, enemy_spawner.enemy_group1, True, False)
        for laser, enemy in collided.items():
            stream.append(str('target_1'))
            tg_1 += 1
            if phase == 1:
                if enemy[0].is_invincible:
                    ##enemy_spawner.enemy_group1.spawn_particles((laser.rect.x, laser.rect.y))
                    bsl_target += 1
                    enemy[0].get_hitp1()
                    enemy[0].spawn_particles((laser.rect.x, laser.rect.y + 11))
                elif not enemy[0].is_invincible:
                    enemy[0].get_hitp1()
                    player.hud.score.update_score(enemy[0].point_value)
                    score += 5
                    stream.append(str('reinforcement'))



        collided = pygame.sprite.groupcollide(player.lasers, enemy_spawner.enemy_group2, True, False)
        for laser, enemy in collided.items():
            alt_1 += 1
            stream.append(str('alt_1'))
            if trial_type == 0:
                sd_alt_1 += 1
            if trial_type == 1:
                sdelta_alt_1 += 1
            if phase == 2:
                if enemy[0].is_invincible:
                    enemy[0].get_hit()
                    enemy[0].spawn_particles((laser.rect.x, laser.rect.y + 11))
                elif not enemy[0].is_invincible:
                    enemy[0].get_hit()
                    player.hud.score.update_score(enemy[0].point_value)
                    score += 5
                    stream.append(str('reinforcement'))
            if phase == 3:
                if trial_type == 0:
                    if enemy[0].is_invincible:
                        enemy[0].get_hit()
                        enemy[0].spawn_particles((laser.rect.x, laser.rect.y + 11))
                    elif not enemy[0].is_invincible:
                        enemy[0].get_hit()
                        player.hud.score.update_score(enemy[0].point_value)
                        score += 5
                        stream.append(8)

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.yasteroid_group, False, True)  ## seperate badguy1 as alt1 group
        for player, new_asteroid in collided.items():
            ##shield_spawner.spawn_shipshield((player.rect.x, player.rect.y))
            new_asteroid[0].get_hit()
            if new_asteroid[0].is_destroyed:
                stream.append(5)
                asteroids += 1
                player.asteroid_hud.yscore.update_score(new_asteroid[0].point_value)
                explosion_spawner.spawn_p1asteroidexplosions((new_asteroid[0].rect.x, new_asteroid[0].rect.y))

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.basteroid_group, False,
                                              True)  ## seperate badguy1 as alt1 group
        for player, new_asteroid in collided.items():
            new_asteroid[0].get_hit()
            stream.append(5)
            asteroids += 1
            player.asteroid_hud.bscore.update_score(new_asteroid[0].point_value)
            explosion_spawner.spawn_p1asteroidexplosions((new_asteroid[0].rect.x, new_asteroid[0].rect.y))

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.rasteroid_group, False,
                                              True)  ## seperate badguy1 as alt1 group
        for player, new_asteroid in collided.items():
            new_asteroid[0].get_hit()
            stream.append(5)
            asteroids += 1
            player.asteroid_hud.rscore.update_score(new_asteroid[0].point_value)
            explosion_spawner.spawn_p1asteroidexplosions((new_asteroid[0].rect.x, new_asteroid[0].rect.y))


        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.enemy_group1, False,
                                              False)  ## seperate badguy1 as alt1 group
        for player, enemy in collided.items():
            player.kickbacked = True
            crash += 1
            if not player.is_invincible:
                player.start_shields()
                if player.is_moving_y:
                    player.vel_y *= -1
                if player.is_moving_x:
                    player.vel_x *= -1


           ## if player.rect.y > enemy[0].rect.y:
             ##   player.rect.y += 100
               ## player.vel_y = 0
            ##elif player.rect.y < enemy[0].rect.y:
              ##  player.rect.y -= 100
                ##player.vel_y = 0
            ##if player.rect.x > enemy[0].rect.x:
              ##  player.rect.x += 100
                ##player.vel_x = 0
            ##if player.rect.x < enemy[0].rect.x:
              ##  player.rect.x -= 100
                ##player.vel_x = 0

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.enemy_group2, False,
                                              False)  ## seperate badguy1 as alt1 group
        for player, enemy in collided.items():
            player.kickbacked = True
            crash += 1
            if not player.is_invincible:
                player.start_shields()
                if player.is_moving_y:
                    player.vel_y *= -1
                if player.is_moving_x:
                    player.vel_x *= -1

            ##if player.rect.y > enemy[0].rect.y:
              ##  player.rect.y += 100
                ##player.vel_y = 0
            ##elif player.rect.y < enemy[0].rect.y:
              ##  player.rect.y -= 100
               ## player.vel_y = 0
            ##if player.rect.x > enemy[0].rect.x:
              ##  player.rect.x += 100
                ##player.vel_x = 0
            ##if player.rect.x < enemy[0].rect.x:
              ##  player.rect.x -= 100
                ##player.vel_x = 0

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.spin_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("spin"))
            spins += 1
            player.turning_left = False
            player.turning_right = False
            player.rolling = False
            player.spinning = True

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.speed_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("speed"))
            turbo += 1
            player.fast = True

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.max_gems_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("coin_boost"))
            coin_boosters += 1
            enemy_spawner.max_generate = True
            enemy_spawner.asteroid_timer = 0

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.roll_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("barrel_roll"))
            barrel_rolls += 1
            if player.turning_left:
                player.rolling_left = True
                player.start_roll_effect()
            elif player.turning_right:
                player.rolling_right = True
                player.start_roll_effect()
            else:
                player.rolling_right = True
                player.start_roll_effect()
            player.turning_left = False
            player.turning_right = False
            player.spinning = False

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.yellow_icon_group, False, True)
        for player, powerup in collided.items():
            yellow_gem_token += 1
            stream.append(str("yellow_tokens"))
            enemy_spawner.blue_gems = False
            enemy_spawner.red_gems = False
            enemy_spawner.yellow_gems = True
            enemy_spawner.icon_timer = 900

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.blue_icon_group, False, True)
        for player, powerup in collided.items():
            blue_gem_token += 1
            stream.append(str("blue_tokens"))
            enemy_spawner.blue_gems = True
            enemy_spawner.red_gems = False
            enemy_spawner.yellow_gems = False
            enemy_spawner.icon_timer = 900

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.red_icon_group, False, True)
        for player, powerup in collided.items():
            red_gem_token += 1
            stream.append(str("red_tokens"))
            enemy_spawner.blue_gems = False
            enemy_spawner.red_gems = True
            enemy_spawner.yellow_gems = False
            enemy_spawner.icon_timer = 900

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.blue_flame_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("blue_flame"))
            flame_change += 1
            player.engine_die()
            player.red_flames = False
            enemy_spawner.red_flames = False
            player.blue_flames = True
            enemy_spawner.blue_flames = True

        collided = pygame.sprite.groupcollide(sprite_group, enemy_spawner.red_flame_group, False, True)
        for player, powerup in collided.items():
            stream.append(str("red_flame"))
            flame_change += 1
            player.engine_die()
            player.blue_flames = False
            enemy_spawner.blue_flames = False
            player.red_flames = True
            enemy_spawner.red_flames = True
            ##collided = pygame.sprite.groupcollide(enemy_spawner.enemy_group1, enemy_spawner.asteroid_group, False, True)  ## seperate badguy1 as alt1 group
            ##for enemy, asteroid in collided.items():
                ##shield_spawner.spawn_shipshield((player.rect.x, player.rect.y))
                ##asteroid[0].get_hit()

                ##if asteroid[0].is_destroyed:
                  ##  explosion_spawner.spawn_p1asteroidexplosions((asteroid[0].rect.x, asteroid[0].rect.y))
            ##for enemy in collided.items():
              ##  enemy[0].new_shield()



def intro():
    display = pygame.display.set_mode(c.display_size)
    black = (0, 0, 0)
    display.fill(black)
    question = 1
    run = True
    ready_text = pygame.image.load("Welcome_Text_01.png")
    question_1 = pygame.image.load("question_1.png")
    question_2 = pygame.image.load("question_2.png")
    bg1 = BG1()
    bg1_group = pygame.sprite.Group()
    bg1_group.add(bg1)
    while run:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                ##if event.key == pygame.K_RETURN:
                    ##main()
                if event.key == pygame.K_SPACE:
                    if question == 1:
                        question = 2
                if event.key == pygame.K_a:
                    if question == 2:
                        question = 1
                    elif question == 3:
                        main()
                if event.key == pygame.K_b:
                    if question == 2:
                        question = 1
                    elif question == 3:
                        question = 1
                if event.key == pygame.K_c:
                    if question == 3:
                        question = 1
                    if question == 2:
                        question = 3
                if event.key == pygame.K_d:
                    if question == 2:
                        question = 1
                    elif phase == 3:
                        question = 1
            display.fill(black)
            if question == 1:
                display.blit(ready_text, (c.display_width // 4, c.display_height // 3))
            elif question == 2:
                display.blit(question_1, (c.display_width // 3, c.display_height // 3))
            elif question == 3:
                display.blit(question_2, (c.display_width // 3, c.display_height // 3))
            pygame.display.update()


        ##bg1_group.update()


def game_over():
    display = pygame.display.set_mode(c.display_size)
    black = (0, 0, 0)
    display.fill(black)
    run = True
    game_over_text = pygame.image.load("gave_over_text.png")
    bg1 = BG1()
    bg1_group = pygame.sprite.Group()
    bg1_group.add(bg1)
    while run:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN:
                    pass
        display.fill(black)
       ##bg1_group.draw(display)  ## draws background with stars
        display.blit(game_over_text, (c.display_width // 2 - 175, c.display_height // 2 - 65))
        pygame.display.update()
        ##bg1_group.update()



##main()

intro()



